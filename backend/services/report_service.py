"""Sales reports (daily / monthly / custom range) as PDF and Excel.

All report types share one aggregator, _aggregate(), which summarises every
bill in a datetime window with optional filters (category, product, bill
number, min/max amount). PDF via reportlab, Excel via openpyxl.
"""
from __future__ import annotations

import io
from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime, time

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import func
from sqlalchemy.orm import Session

from database.crud import repositories as repo
from backend.services.timezone_util import ist_date_str, ist_time_str, ist_now_str
from database.models import Bill, BillItem, Product

BRAND = colors.HexColor("#0f766e")
LIGHT = colors.HexColor("#e6f2f0")


def _rupees(v: float) -> str:
    return f"Rs.{(v or 0):,.2f}"


def _aggregate(session: Session, start: datetime, end: datetime, *,
               category_id: int | None = None, product_name: str | None = None,
               bill_number: str | None = None, min_amount: float | None = None,
               max_amount: float | None = None) -> dict:
    """Summarise all bills in [start, end] with optional filters."""
    q = session.query(Bill).filter(Bill.bill_date >= start, Bill.bill_date <= end)
    if bill_number:
        q = q.filter(Bill.bill_number.ilike(f"%{bill_number.strip()}%"))
    if min_amount is not None:
        q = q.filter(Bill.grand_total >= min_amount)
    if max_amount is not None:
        q = q.filter(Bill.grand_total <= max_amount)
    bills = q.order_by(Bill.bill_date.asc()).all()

    # Category / product filters act on bill contents: keep bills that contain
    # at least one matching line.
    if category_id is not None or product_name:
        keep = []
        for b in bills:
            match = False
            for it in b.items:
                if it.product_id is None:
                    continue
                p = repo.products.get(session, it.product_id)
                if not p:
                    continue
                if category_id is not None and p.category_id != category_id:
                    continue
                if product_name and product_name.strip().lower() not in (p.product_name or "").lower():
                    continue
                match = True
                break
            if match:
                keep.append(b)
        bills = keep

    total_bills = len(bills)
    total_items = sum(b.total_items or 0 for b in bills)
    gross = sum(b.subtotal or 0 for b in bills)
    discount = sum(b.total_discount or 0 for b in bills)
    net = sum(b.grand_total or 0 for b in bills)

    pay = defaultdict(float)
    for b in bills:
        method = (b.payment_method or "cash").lower()
        if method == "split" and getattr(b, "payment_breakdown", None):
            import json as _json
            try:
                parts = _json.loads(b.payment_breakdown)
                for k in ("cash", "upi", "card"):
                    pay[k] += float(parts.get(k, 0) or 0)
            except (ValueError, TypeError):
                pay["cash"] += b.grand_total or 0
        else:
            pay[method] += b.grand_total or 0
    # Cash vs online split for the end-of-day summary.
    cash_sales = round(pay.get("cash", 0), 2)
    online_sales = round(pay.get("upi", 0) + pay.get("card", 0), 2)

    # Product-wise totals (qty + revenue) across the kept bills.
    prod_qty: dict[str, int] = defaultdict(int)
    prod_rev: dict[str, float] = defaultdict(float)
    for b in bills:
        for it in b.items:
            if it.product_id is not None:
                p = repo.products.get(session, it.product_id)
                name = p.product_name if p else "(deleted product)"
            else:
                name = it.item_name or "(manual item)"
            prod_qty[name] += it.quantity or 0
            prod_rev[name] += it.total_price or 0
    top = sorted(prod_qty.keys(), key=lambda n: prod_qty[n], reverse=True)
    top_products = [(n, prod_qty[n], prod_rev[n]) for n in top]

    # Daily breakdown (for monthly/range reports).
    by_day: dict[date, float] = defaultdict(float)
    for b in bills:
        by_day[b.bill_date.date()] += b.grand_total or 0
    daily = sorted(by_day.items())

    return {
        "bills": bills,
        "total_bills": total_bills, "total_items": total_items,
        "gross": gross, "discount": discount, "net": net,
        "pay": pay, "top_products": top_products, "daily": daily,
        "cash_sales": cash_sales, "online_sales": online_sales,
    }


# --------------------------------------------------------------------------- PDF
def _styles():
    base = getSampleStyleSheet()
    return {
        "brand": ParagraphStyle("brand", parent=base["Title"], textColor=BRAND, fontSize=20, spaceAfter=2),
        "sub": ParagraphStyle("sub", parent=base["Normal"], fontSize=12, textColor=colors.HexColor("#444")),
        "sec": ParagraphStyle("sec", parent=base["Heading2"], fontSize=13, textColor=BRAND, spaceBefore=12, spaceAfter=6),
        "small": ParagraphStyle("small", parent=base["Normal"], fontSize=9, textColor=colors.HexColor("#666")),
    }


def _header_table(rows, col_widths, aligns=None):
    t = Table(rows, colWidths=col_widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), BRAND),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d0d8d6")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    for col, al in (aligns or {}).items():
        style.append(("ALIGN", (col, 0), (col, -1), al))
    t.setStyle(TableStyle(style))
    return t


def build_report_pdf(data: dict, *, report_type: str, period_label: str,
                     include_daily: bool = False, cash: dict | None = None) -> bytes:
    S = _styles()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=16 * mm, bottomMargin=16 * mm,
                            leftMargin=16 * mm, rightMargin=16 * mm, title=f"{report_type} {period_label}")
    story = [
        Paragraph("RESVI ENTERPRISES", S["brand"]),
        Paragraph(report_type, S["sub"]),
        Paragraph(f"{period_label} &nbsp;&nbsp;|&nbsp;&nbsp; Generated: "
                  f"{ist_now_str()}", S["small"]),
        Spacer(1, 8),
        Paragraph("Summary", S["sec"]),
    ]
    summary = [
        ["Total Bills", str(data["total_bills"])],
        ["Total Items Sold", str(data["total_items"])],
        ["Gross Sales", _rupees(data["gross"])],
        ["Total Discount", _rupees(data["discount"])],
        ["Net Sales", _rupees(data["net"])],
    ]
    st = Table(summary, colWidths=[70 * mm, 100 * mm])
    st.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), LIGHT),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (0, 4), (-1, 4), "Helvetica-Bold"),
        ("TEXTCOLOR", (1, 4), (1, 4), BRAND),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d0d8d6")),
    ]))
    story.append(st)

    if include_daily and data["daily"]:
        story.append(Paragraph("Daily Sales Breakdown", S["sec"]))
        rows = [["Date", "Net Sales"]]
        rows += [[d.strftime("%d %b"), _rupees(v)] for d, v in data["daily"]]
        story.append(_header_table(rows, [90 * mm, 80 * mm], {1: "RIGHT"}))

    story.append(Paragraph("Top Selling Products", S["sec"]))
    if data["top_products"]:
        rows = [["Product Name", "Qty Sold", "Revenue"]]
        rows += [[n, str(q), _rupees(r)] for n, q, r in data["top_products"][:10]]
        story.append(_header_table(rows, [90 * mm, 35 * mm, 45 * mm], {1: "RIGHT", 2: "RIGHT"}))
    else:
        story.append(Paragraph("No products sold in this period.", S["small"]))

    story.append(Paragraph("Bill-wise Summary", S["sec"]))
    if data["bills"]:
        rows = [["Bill Number", "Date", "Time", "Items", "Amount"]]
        for b in data["bills"]:
            rows.append([
                b.bill_number,
                ist_date_str(b.bill_date),
                ist_time_str(b.bill_date),
                str(b.total_items or 0),
                _rupees(b.grand_total),
            ])
        story.append(_header_table(rows, [42 * mm, 32 * mm, 30 * mm, 22 * mm, 44 * mm],
                                   {3: "RIGHT", 4: "RIGHT"}))
    else:
        story.append(Paragraph("No bills in this period.", S["small"]))

    story.append(Paragraph("Payment Summary", S["sec"]))
    prows = [["Method", "Amount"]]
    for m in ("cash", "upi", "card"):
        prows.append([m.upper(), _rupees(data["pay"].get(m, 0))])
    prows.append(["ONLINE (UPI + Card)", _rupees(data.get("online_sales", 0))])
    prows.append(["CASH", _rupees(data.get("cash_sales", 0))])
    prows.append(["TOTAL", _rupees(data["net"])])
    story.append(_header_table(prows, [70 * mm, 100 * mm], {1: "RIGHT"}))

    # Cash drawer section (daily reports only).
    if cash:
        story.append(Paragraph("Cash Drawer", S["sec"]))
        crows = [["Item", "Amount"]]
        crows.append(["Opening Cash", _rupees(cash.get("opening_cash", 0))])
        crows.append(["Cash Sales", _rupees(cash.get("cash_sales", 0))])
        crows.append(["Cash Expenses", _rupees(cash.get("cash_expenses", 0))])
        crows.append(["Expected Cash", _rupees(cash.get("expected_cash", 0))])
        if cash.get("actual_cash") is not None:
            crows.append(["Actual Cash", _rupees(cash.get("actual_cash", 0))])
            crows.append(["Difference", _rupees(cash.get("difference", 0))])
            crows.append(["Closing Cash", _rupees(cash.get("closing_cash", 0))])
        story.append(_header_table(crows, [70 * mm, 100 * mm], {1: "RIGHT"}))

        # Itemized expense list.
        exp_list = cash.get("expenses") or []
        if exp_list:
            story.append(Paragraph("Expense Summary", S["sec"]))
            erows = [["Description", "Amount"]]
            for e in exp_list:
                erows.append([e.get("description", ""), _rupees(e.get("amount", 0))])
            erows.append(["Total Expenses", _rupees(cash.get("cash_expenses", 0))])
            story.append(_header_table(erows, [120 * mm, 50 * mm], {1: "RIGHT"}))

    story.append(Spacer(1, 18))
    story.append(Paragraph("Generated by RESVI Billing App", S["small"]))
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


# ------------------------------------------------------------------------- Excel
def build_report_excel(data: dict, *, report_type: str, period_label: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="0F766E")
    head_font = Font(bold=True, color="FFFFFF")

    ws = wb.active
    ws.title = "Summary"
    ws.append(["RESVI ENTERPRISES"])
    ws.append([report_type, period_label])
    ws.append([])
    for k, v in [("Total Bills", data["total_bills"]), ("Total Items Sold", data["total_items"]),
                 ("Gross Sales", round(data["gross"], 2)), ("Total Discount", round(data["discount"], 2)),
                 ("Net Sales", round(data["net"], 2))]:
        ws.append([k, v])

    wb_bills = wb.create_sheet("Bills")
    wb_bills.append(["Bill Number", "Date", "Time", "Items", "Amount"])
    for c in wb_bills[1]:
        c.fill = head_fill; c.font = head_font
    for b in data["bills"]:
        wb_bills.append([
            b.bill_number,
            ist_date_str(b.bill_date),
            ist_time_str(b.bill_date),
            b.total_items or 0, round(b.grand_total or 0, 2),
        ])

    wb_prod = wb.create_sheet("Products")
    wb_prod.append(["Product Name", "Qty Sold", "Revenue"])
    for c in wb_prod[1]:
        c.fill = head_fill; c.font = head_font
    for n, q, r in data["top_products"]:
        wb_prod.append([n, q, round(r, 2)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ------------------------------------------------------------------- range helpers
def day_range(d: date) -> tuple[datetime, datetime]:
    return datetime.combine(d, time.min), datetime.combine(d, time.max)


def month_range(year: int, month: int) -> tuple[datetime, datetime]:
    last = monthrange(year, month)[1]
    return (datetime.combine(date(year, month, 1), time.min),
            datetime.combine(date(year, month, last), time.max))
