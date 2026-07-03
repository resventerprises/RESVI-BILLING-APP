"""
Bulk product import.

Reads a products.xlsx with columns (exact order, header row required):
    barcode | product_name | price | quantity | category | image_name

Optionally takes a ZIP of product images; each row's `image_name` is matched to
a file inside the ZIP, saved under uploads/products/, and linked to the product.

Behaviour:
  * Categories are created on demand (matched case-insensitively by name).
  * A row whose barcode already exists is UPDATED (price/qty/name/category/image),
    not duplicated — so re-importing a corrected sheet is safe.
  * Rows with a blank barcode get an auto-generated one.
  * Every row's outcome is reported back (created / updated / skipped + reason)
    so the user sees exactly what happened.
"""
from __future__ import annotations

import io
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.services import category_service
from backend.services.code_generator import next_barcode, next_product_code
from config import settings
from database.crud import repositories as repo
from database.models import Product, Status

EXPECTED_COLUMNS = ["barcode", "product_name", "price", "quantity", "category", "image_name"]

# Header aliases -> canonical field. Order-independent; supports both the
# "image_name" and "image file" / "min stock" spec variants.
COLUMN_ALIASES = {
    "barcode": "barcode",
    "product_name": "product_name", "product name": "product_name", "name": "product_name",
    "price": "price", "selling_price": "price", "selling price": "price",
    "quantity": "quantity", "qty": "quantity", "stock": "quantity",
    "category": "category",
    "image_name": "image_name", "image name": "image_name", "image": "image_name",
    "image file": "image_name", "image_file": "image_name",
    "min_stock": "min_stock", "min stock": "min_stock", "minimum stock": "min_stock",
    "min_stock_level": "min_stock", "minimum_stock": "min_stock",
    "description": "description",
}


def _map_header(header_row) -> dict[str, int]:
    """Map canonical field -> column index, from a header row (any order)."""
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        key = _norm(raw).lower()
        field = COLUMN_ALIASES.get(key)
        if field and field not in mapping:
            mapping[field] = idx
    return mapping


def _norm(v) -> str:
    return ("" if v is None else str(v)).strip()


def _product_dir(product_id: int) -> Path:
    d = Path(settings.UPLOAD_DIR) / str(product_id)
    d.mkdir(parents=True, exist_ok=True)
    return d


def _load_zip(zip_bytes: bytes | None) -> dict[str, bytes]:
    """Return {lowercased basename: raw bytes} for every image in the zip."""
    if not zip_bytes:
        return {}
    images: dict[str, bytes] = {}
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            name = Path(info.filename).name  # ignore folder structure inside the zip
            if not name or name.startswith("."):
                continue
            if Path(name).suffix.lower() not in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
                continue
            images[name.lower()] = zf.read(info)
    return images


def _category_id(session: Session, cache: dict[str, int], name: str) -> int | None:
    name = _norm(name)
    if not name:
        return None
    key = name.lower()
    if key in cache:
        return cache[key]
    # Match existing (case-insensitive) or create.
    for c in category_service.list_categories(session):
        if c.category_name.strip().lower() == key:
            cache[key] = c.id
            return c.id
    created = category_service.create_category(session, name)
    session.flush()
    cache[key] = created.id
    return created.id


def _read_rows(file_bytes: bytes, filename: str) -> list[list]:
    """Return all rows (including header) from an .xlsx or .csv file."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        import csv as _csv

        text = file_bytes.decode("utf-8-sig", errors="replace")
        return [list(r) for r in _csv.reader(io.StringIO(text))]
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    return [list(r) for r in ws.iter_rows(values_only=True)]


def import_products(session: Session, recognizer, file_bytes: bytes,
                    zip_bytes: bytes | None = None, filename: str = "products.xlsx") -> dict:
    rows = _read_rows(file_bytes, filename)
    if not rows:
        return {"created": 0, "updated": 0, "skipped": 0, "failed": 0,
                "errors": ["The file is empty."], "rows": []}

    colmap = _map_header(rows[0])
    if "product_name" not in colmap or "price" not in colmap:
        return {
            "created": 0, "updated": 0, "skipped": 0, "failed": 0,
            "errors": ["Header must include at least: product_name and price. "
                       "Recognised columns: product_name, price, quantity, category, barcode, "
                       "min_stock, image_name."],
            "rows": [],
        }

    def cell(row, field):
        idx = colmap.get(field)
        return _norm(row[idx]) if idx is not None and idx < len(row) else ""

    images = _load_zip(zip_bytes)
    # Also allow images already sitting in uploads/products/ (spec point 5).
    disk_images: dict[str, Path] = {}
    updir = Path(settings.UPLOAD_DIR)
    if updir.is_dir():
        for f in updir.rglob("*"):
            if f.is_file() and f.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
                disk_images.setdefault(f.name.lower(), f)

    cat_cache: dict[str, int] = {}
    seen_barcodes: set[str] = set()
    created = updated = skipped = failed = 0
    report: list[dict] = []

    for i, raw in enumerate(rows[1:], start=2):
        if raw is None or all(_norm(c) == "" for c in raw):
            continue  # blank line
        barcode = cell(raw, "barcode")
        name = cell(raw, "product_name")
        entry = {"row": i, "barcode": barcode, "product_name": name}
        try:
            # ---- Validation -------------------------------------------------
            if not name:
                raise ValueError("Product Name is required")
            try:
                price_val = float(cell(raw, "price") or 0)
            except ValueError:
                raise ValueError("Price must be a number")
            if price_val <= 0:
                raise ValueError("Price must be greater than 0")
            qty_raw = cell(raw, "quantity")
            try:
                qty_val = int(float(qty_raw or 0))
            except ValueError:
                raise ValueError("Quantity must be a whole number")
            if qty_val < 0:
                raise ValueError("Quantity cannot be negative")
            min_raw = cell(raw, "min_stock")
            min_stock = int(float(min_raw)) if min_raw else 0
            if barcode and barcode in seen_barcodes:
                raise ValueError(f"Duplicate barcode {barcode} within this file")
            if barcode:
                seen_barcodes.add(barcode)

            cat_id = _category_id(session, cat_cache, cell(raw, "category")) \
                or _category_id(session, cat_cache, "Uncategorized")

            existing = session.query(Product).filter(Product.barcode == barcode).first() if barcode else None

            if existing:
                existing.product_name = name
                existing.selling_price = price_val
                existing.category_id = cat_id
                if min_stock:
                    existing.min_stock_level = min_stock
                if qty_val and (existing.quantity or 0) != qty_val:
                    from backend.services import inventory_service
                    inventory_service.adjust(session, existing.id, qty_val - (existing.quantity or 0),
                                             remarks="Bulk import update")
                product = existing
                updated += 1
                entry["status"] = "updated"
            else:
                if not barcode:
                    barcode = next_barcode(session)
                    entry["barcode"] = barcode
                product = repo.products.create(
                    session,
                    product_code=next_product_code(session),
                    barcode=barcode,
                    product_name=name,
                    category_id=cat_id,
                    selling_price=price_val,
                    discount=0.0,
                    cost_price=0.0,
                    quantity=0,
                    min_stock_level=min_stock,
                    description=cell(raw, "description") or None,
                    family_key=None,
                    status=Status.ACTIVE,
                )
                session.flush()
                if qty_val > 0:
                    from backend.services import inventory_service
                    inventory_service.stock_in(session, product.id, qty_val, remarks="Bulk import opening stock")
                created += 1
                entry["status"] = "created"

            # ---- Image mapping: from zip first, else uploads/products/ ------
            image_name = cell(raw, "image_name")
            if image_name:
                raw_img = images.get(image_name.lower())
                if raw_img is not None:
                    ext = Path(image_name).suffix.lower() or ".jpg"
                    dest = _product_dir(product.id) / f"{product.id}_{len(repo.product_images.for_product(session, product.id))}{ext}"
                    dest.write_bytes(raw_img)
                    repo.product_images.create(session, product_id=product.id, image_path=str(dest), image_type="import")
                    entry["image"] = "attached"
                elif image_name.lower() in disk_images:
                    src = disk_images[image_name.lower()]
                    repo.product_images.create(session, product_id=product.id, image_path=str(src), image_type="import")
                    entry["image"] = "linked"
                else:
                    entry["image"] = "not found"

            report.append(entry)
        except Exception as exc:  # noqa: BLE001 - per-row; keep going
            failed += 1
            entry["status"] = "failed"
            entry["error"] = str(exc)
            report.append(entry)

    return {"created": created, "updated": updated, "skipped": skipped,
            "failed": failed, "errors": [], "rows": report}
