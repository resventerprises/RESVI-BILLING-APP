"""Export the full product catalogue to Excel.

Exports EVERY product in the database (imported + manually added), not just the
ones from the original spreadsheet. The header names are chosen so the file can
be edited and fed straight back into Import Products: the importer maps
"Product Name" -> product_name, "Minimum Stock" -> min_stock, etc. The extra
columns (Serial No, Status, dates) are ignored by the importer.
"""
from __future__ import annotations

import io

from sqlalchemy import select
from sqlalchemy.orm import Session

from backend.services.timezone_util import ist_date_str
from database.models import Category, Product, Status

HEADERS = [
    "Serial No", "Product Name", "Barcode", "Category", "Price",
    "Quantity", "Minimum Stock", "Status", "Created Date", "Last Updated",
]


def build_products_excel(session: Session, include_inactive: bool = False) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    stmt = select(Product)
    if not include_inactive:
        stmt = stmt.where(Product.status == Status.ACTIVE)
    products = session.scalars(stmt.order_by(Product.product_name)).all()

    cats = {c.id: c.category_name for c in session.scalars(select(Category)).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    head_fill = PatternFill("solid", fgColor="0F766E")
    head_font = Font(bold=True, color="FFFFFF")
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, p in enumerate(products, start=1):
        status = p.status.value if hasattr(p.status, "value") else str(p.status)
        ws.append([
            i,
            p.product_name or "",
            p.barcode or "",
            cats.get(p.category_id, "") or "",
            round(float(p.selling_price or 0), 2),
            int(p.quantity or 0),
            int(p.min_stock_level or 0),
            status,
            ist_date_str(p.created_at) if getattr(p, "created_at", None) else "",
            ist_date_str(p.updated_at) if getattr(p, "updated_at", None) else "",
        ])

    widths = [10, 42, 18, 20, 12, 11, 15, 12, 14, 14]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
